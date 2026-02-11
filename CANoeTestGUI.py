"""
CANoe Automation Test Tool - Professional GUI
==============================================
PyQt6-based GUI for CANoe Software-in-the-Loop (SIL) testing.
Wraps the CANoe COM API with proper threading architecture
to keep the UI responsive during long COM operations.

Architecture:
    UI Thread (MainWindow)  <--signals/slots-->  Worker Thread (CANoeWorker)
                                                    |
                                                    v
                                              CANoe COM Server
"""

import sys
import os
import time
import datetime
import traceback
import subprocess

# ---------------------------------------------------------------------------
# Dependency checks
# ---------------------------------------------------------------------------
try:
    from PyQt6.QtWidgets import (
        QApplication,
        QMainWindow,
        QWidget,
        QVBoxLayout,
        QHBoxLayout,
        QGroupBox,
        QLabel,
        QLineEdit,
        QPushButton,
        QTextEdit,
        QTreeWidget,
        QTreeWidgetItem,
        QProgressBar,
        QSplitter,
        QFileDialog,
        QMessageBox,
        QHeaderView,
        QSizePolicy,
        QCheckBox,
        QSpinBox,
        QDateTimeEdit,
        QAbstractItemView,
        QMenu,
    )
    from PyQt6.QtCore import (
        Qt,
        QThread,
        QObject,
        pyqtSignal,
        pyqtSlot,
        QTimer,
        QSettings,
        QDateTime,
    )
    from PyQt6.QtGui import (
        QFont,
        QColor,
        QTextCursor,
        QAction,
        QPixmap,
        QIcon,
    )
except ImportError:
    print("ERROR: PyQt6 is required.  pip install PyQt6")
    sys.exit(1)

try:
    import pythoncom
    from win32com.client import DispatchEx, CastTo, WithEvents, DispatchWithEvents

    HAS_COM = True
except ImportError:
    HAS_COM = False
    print("WARNING: pywin32 not found. COM features will be disabled.")

try:
    import openpyxl
    from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
APP_TITLE = "BLTN Team - YuRa HaNoi "
VERSION = "2.0.0"

VERDICT_NA = 0
VERDICT_PASSED = 1
VERDICT_FAILED = 2

LED_COLORS = {
    "idle": "#6c6c6c",
    "running": "#569cd6",
    "pass": "#4caf50",
    "fail": "#f44336",
    "error": "#ff9800",
}


# Resolve resource path: works both in dev and when packaged by PyInstaller
if getattr(sys, "frozen", False):
    _SCRIPT_DIR = sys._MEIPASS
else:
    _SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(_SCRIPT_DIR, "YuRa.png")


# ---------------------------------------------------------------------------
# Dark professional stylesheet (automotive-tool look)
# ---------------------------------------------------------------------------
DARK_STYLE = """
QMainWindow, QWidget {
    background-color: #1e1e1e;
    color: #d4d4d4;
    font-family: "Segoe UI", sans-serif;
    font-size: 10pt;
}
QGroupBox {
    border: 1px solid #3c3c3c;
    border-radius: 4px;
    margin-top: 14px;
    padding-top: 14px;
    font-weight: bold;
    color: #569cd6;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 6px;
}
QLineEdit {
    background-color: #2d2d2d;
    border: 1px solid #3c3c3c;
    border-radius: 3px;
    padding: 5px 8px;
    color: #d4d4d4;
    selection-background-color: #264f78;
}
QLineEdit:focus { border-color: #569cd6; }

QPushButton {
    background-color: #0e639c;
    color: #ffffff;
    border: none;
    border-radius: 3px;
    padding: 6px 16px;
    font-weight: bold;
    min-height: 24px;
}
QPushButton:hover   { background-color: #1177bb; }
QPushButton:pressed { background-color: #0d5689; }
QPushButton:disabled {
    background-color: #2a2a2a;
    color: #555555;
    font-weight: normal;
}
QPushButton#btn_start       { background-color: #388e3c; color: #ffffff; }
QPushButton#btn_start:hover { background-color: #43a047; color: #ffffff; }
QPushButton#btn_start:disabled { background-color: #2a3a2a; color: #4a5a4a; }

QPushButton#btn_stop        { background-color: #d32f2f; color: #ffffff; }
QPushButton#btn_stop:hover  { background-color: #e53935; color: #ffffff; }
QPushButton#btn_stop:disabled { background-color: #3a2a2a; color: #5a4a4a; }

QPushButton#btn_run         { background-color: #f57c00; color: #ffffff; }
QPushButton#btn_run:hover   { background-color: #fb8c00; color: #ffffff; }
QPushButton#btn_run:disabled { background-color: #3a3a2a; color: #5a5a4a; }

QTextEdit {
    background-color: #1e1e1e;
    border: 1px solid #3c3c3c;
    border-radius: 3px;
    color: #d4d4d4;
    font-family: "Consolas", "Courier New", monospace;
    font-size: 9pt;
}
QTreeWidget {
    background-color: #252526;
    border: 1px solid #3c3c3c;
    border-radius: 3px;
    color: #d4d4d4;
    alternate-background-color: #2a2d2e;
}
QTreeWidget::item:hover    { background-color: #2a2d2e; }
QTreeWidget::item:selected { background-color: #094771; }
QTreeWidget::indicator:unchecked {
    border: 1px solid #555;
    background-color: #2d2d2d;
}
QTreeWidget::indicator:checked {
    border: 1px solid #569cd6;
    background-color: #569cd6;
}
QHeaderView::section {
    background-color: #252526;
    color: #d4d4d4;
    padding: 4px;
    border: 1px solid #3c3c3c;
}
QProgressBar {
    border: 1px solid #3c3c3c;
    border-radius: 3px;
    background-color: #2d2d2d;
    text-align: center;
    color: #d4d4d4;
    height: 22px;
}
QProgressBar::chunk {
    background-color: #569cd6;
    border-radius: 2px;
}
QStatusBar {
    background-color: #007acc;
    color: #ffffff;
    font-size: 9pt;
}
QStatusBar::item { border: none; }
QMenuBar {
    background-color: #2d2d2d;
    color: #d4d4d4;
    border-bottom: 1px solid #3c3c3c;
}
QMenuBar::item:selected { background-color: #094771; }
QMenu {
    background-color: #252526;
    color: #d4d4d4;
    border: 1px solid #3c3c3c;
}
QMenu::item:selected { background-color: #094771; }
QSplitter::handle {
    background-color: #3c3c3c;
    width: 2px;
}
QSpinBox {
    background-color: #2d2d2d;
    border: 1px solid #3c3c3c;
    border-radius: 3px;
    padding: 3px 6px;
    color: #d4d4d4;
    min-height: 22px;
}
QSpinBox:focus { border-color: #569cd6; }
QSpinBox::up-button, QSpinBox::down-button {
    background-color: #3c3c3c;
    border: none;
    width: 16px;
}
QSpinBox::up-button:hover, QSpinBox::down-button:hover {
    background-color: #505050;
}
QDateTimeEdit {
    background-color: #2d2d2d;
    border: 1px solid #3c3c3c;
    border-radius: 3px;
    padding: 3px 6px;
    color: #d4d4d4;
    min-height: 22px;
}
QDateTimeEdit:focus { border-color: #569cd6; }
QDateTimeEdit::up-button, QDateTimeEdit::down-button {
    background-color: #3c3c3c;
    border: none;
    width: 16px;
}
QDateTimeEdit::up-button:hover, QDateTimeEdit::down-button:hover {
    background-color: #505050;
}
QCheckBox {
    spacing: 6px;
    color: #d4d4d4;
}
QCheckBox::indicator {
    width: 16px;
    height: 16px;
    border: 1px solid #555;
    border-radius: 2px;
    background-color: #2d2d2d;
}
QCheckBox::indicator:checked {
    background-color: #569cd6;
    border-color: #569cd6;
}
QCheckBox::indicator:hover {
    border-color: #569cd6;
}
"""


# ==========================================================================
# COM Event Handlers  (instantiated in the worker thread)
# ==========================================================================


class AppEventHandler:
    """Handles CANoe.Application events."""

    def __init__(self):
        self.worker = None

    def OnOpen(self, fullname):
        if self.worker:
            self.worker.log_message.emit(f"Configuration opened: {fullname}", "info")

    def OnQuit(self):
        if self.worker:
            self.worker.log_message.emit("CANoe application quit", "warning")
            self.worker.canoe_quit.emit()


class MeasEventHandler:
    """Handles CANoe.Measurement events."""

    def __init__(self):
        self.worker = None
        self._started = False
        self._stopped = False

    @property
    def is_started(self):
        return self._started

    @property
    def is_stopped(self):
        return self._stopped

    def OnInit(self):
        if self.worker:
            self.worker.log_message.emit("Measurement initializing...", "info")

    def OnStart(self):
        self._started = True
        self._stopped = False
        if self.worker:
            self.worker.log_message.emit("Measurement started", "info")
            self.worker.measurement_state.emit(True)

    def OnStop(self):
        self._started = False
        self._stopped = True
        if self.worker:
            self.worker.log_message.emit("Measurement stopped", "info")
            self.worker.measurement_state.emit(False)

    def OnExit(self):
        if self.worker:
            self.worker.log_message.emit("Measurement exited", "info")


class TestEventHandler:
    """Handles CANoe TestModule / TestConfiguration events."""

    def __init__(self):
        self.worker = None
        self.started = False
        self.stopped = False
        self.verdict = VERDICT_NA

    def OnStart(self):
        self.started = True
        self.stopped = False
        self.verdict = VERDICT_NA
        if self.worker:
            self.worker.test_item_started.emit(self.Name)

    def OnStop(self, reason):
        self.started = False
        self.stopped = True
        self.verdict = reason
        if self.worker:
            tag = {0: "N/A", 1: "PASSED", 2: "FAILED"}.get(reason, f"Unknown({reason})")
            lvl = "info" if reason == 1 else ("error" if reason == 2 else "warning")
            self.worker.log_message.emit(
                f"Test '{self.Name}' stopped  -  Verdict: {tag}", lvl
            )
            self.worker.test_item_finished.emit(self.Name, reason)


# ==========================================================================
# IntraGroupTreeWidget  (drag-drop reordering within groups only)
# ==========================================================================


class IntraGroupTreeWidget(QTreeWidget):
    """QTreeWidget that only allows drag-drop reordering within the same parent group."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)

    def dropEvent(self, event):
        dragged = self.currentItem()
        if dragged is None:
            event.ignore()
            return
        drop_target = self.itemAt(event.position().toPoint())
        if drop_target is None:
            event.ignore()
            return
        dragged_parent = dragged.parent()
        target_parent = drop_target.parent()
        if target_parent is None:
            # drop_target is a top-level group item
            if drop_target == dragged_parent:
                super().dropEvent(event)
            else:
                event.ignore()
        elif dragged_parent == target_parent:
            super().dropEvent(event)
        else:
            event.ignore()


# ==========================================================================
# CANoe Worker  (lives in a dedicated QThread)
# ==========================================================================


class CANoeWorker(QObject):
    """All CANoe COM operations happen here, off the UI thread."""

    # --- signals to UI ---
    log_message = pyqtSignal(str, str)  # (msg, level)
    state_changed = pyqtSignal(str)  # state name
    canoe_connected = pyqtSignal(str)  # version string
    canoe_quit = pyqtSignal()
    measurement_state = pyqtSignal(bool)  # running?
    test_modules_found = pyqtSignal(list)  # [(name, enabled), ...]
    test_configs_found = pyqtSignal(list)
    test_item_started = pyqtSignal(str)  # name
    test_item_finished = pyqtSignal(str, int)  # name, verdict
    all_tests_complete = pyqtSignal(int, int)  # passed, failed
    progress = pyqtSignal(int, int)  # done, total
    verdict_update = pyqtSignal(str)  # led state key
    report_available = pyqtSignal(str)  # path (global)
    test_report_found = pyqtSignal(str, str)  # test_name, report_path
    write_window_text = pyqtSignal(str)  # new Write Window content
    capl_path_found = pyqtSignal(str, str)  # (name, path)
    test_retry = pyqtSignal(str, int, int)  # (name, attempt, max)
    loop_progress = pyqtSignal(int, int)  # (current_loop, total_loops)

    def __init__(self):
        super().__init__()
        self._app_com = None
        self._meas_com = None
        self._meas_ev = None
        self._app_ev = None
        self._test_modules = []
        self._test_configs = []
        self._config_dir = ""
        self._pump_timer = None
        self._write_poll_timer = None
        self._last_write_text = ""
        self._connected = False

    # ---- thread bootstrap ------------------------------------------------

    @pyqtSlot()
    def initialize(self):
        """Called once when the thread starts."""
        if not HAS_COM:
            self.log_message.emit("pywin32 not installed - COM disabled", "error")
            return
        try:
            pythoncom.CoInitialize()
            self._pump_timer = QTimer()
            self._pump_timer.timeout.connect(self._pump)
            self._pump_timer.start(50)
            self._write_poll_timer = QTimer()
            self._write_poll_timer.timeout.connect(self._poll_write_window)
            self.log_message.emit("COM initialised. Ready.", "info")
        except Exception as e:
            self.log_message.emit(f"COM init failed: {e}", "error")

    @pyqtSlot()
    def _pump(self):
        try:
            pythoncom.PumpWaitingMessages()
        except Exception:
            pass

    @pyqtSlot()
    def _poll_write_window(self):
        """Read CANoe Write Window and emit new content."""
        if not self._connected or not self._app_com:
            return
        try:
            text = self._app_com.UI.Write.Text
            if text is None:
                text = ""
            if text != self._last_write_text:
                if text.startswith(self._last_write_text):
                    new_part = text[len(self._last_write_text) :]
                else:
                    new_part = text
                if new_part.strip():
                    self.write_window_text.emit(new_part)
                self._last_write_text = text
        except Exception:
            pass

    @pyqtSlot()
    def reset_write_cache(self):
        """Reset cached text so next poll re-syncs from current state."""
        self._last_write_text = ""

    # ---- helpers ---------------------------------------------------------

    def _do_events(self):
        pythoncom.PumpWaitingMessages()
        time.sleep(0.1)

    def _wait_for(self, cond, timeout=30):
        t0 = time.time()
        while not cond():
            if time.time() - t0 > timeout:
                return False
            self._do_events()
        return True

    # ---- public slots (called via signals from UI) -----------------------

    @pyqtSlot(str)
    def connect_and_load(self, cfg_path):
        if not HAS_COM:
            self.log_message.emit("COM unavailable", "error")
            return
        try:
            self.state_changed.emit("connecting")
            self.log_message.emit("Connecting to CANoe COM Server...", "info")

            self._app_com = DispatchEx("CANoe.Application")
            ver = self._app_com.Version
            ver_str = f"{ver.major}.{ver.minor}.{ver.Build}"
            self.log_message.emit(f"CANoe {ver_str} loaded", "info")

            # event handlers
            self._app_ev = WithEvents(self._app_com, AppEventHandler)
            self._app_ev.worker = self
            self._meas_com = self._app_com.Measurement
            self._meas_ev = WithEvents(self._meas_com, MeasEventHandler)
            self._meas_ev.worker = self

            # open config
            self._app_com.Configuration.Modified = False
            cfg_abs = os.path.abspath(cfg_path)
            self._config_dir = os.path.dirname(cfg_abs)
            self.log_message.emit(f"Opening: {cfg_abs}", "info")
            self._app_com.Open(cfg_abs)

            self._connected = True
            self._test_modules.clear()
            self._test_configs.clear()
            self._last_write_text = ""
            # Start Write Window polling (500ms interval)
            if self._write_poll_timer and not self._write_poll_timer.isActive():
                self._write_poll_timer.start(500)
            self.canoe_connected.emit(ver_str)
            self.state_changed.emit("connected")
            self.log_message.emit("Configuration loaded successfully", "info")

            # Auto-scan test environments already loaded in CANoe
            self._scan_existing_test_envs()

        except Exception as e:
            self.log_message.emit(f"Connection failed: {e}", "error")
            self.state_changed.emit("error")
            self._connected = False

    def _scan_existing_test_envs(self):
        """Scan CANoe for already-loaded test environments and populate tree."""
        try:
            ts = self._app_com.Configuration.TestSetup
            envs = ts.TestEnvironments
            if envs.Count == 0:
                return
            self._test_modules.clear()
            for i in range(1, envs.Count + 1):
                try:
                    tenv = envs.Item(i)
                    tenv = CastTo(tenv, "ITestEnvironment2")
                    self._collect_modules(tenv)
                except Exception:
                    pass

            # Pump and refresh names
            for _ in range(30):
                self._do_events()
            self._refresh_module_names()

            if self._test_modules:
                info = [(m["name"], m["enabled"]) for m in self._test_modules]
                self.test_modules_found.emit(info)
                self.log_message.emit(
                    f"Found {len(self._test_modules)} existing test module(s) - all enabled",
                    "info",
                )
                # Auto-scan for existing reports
                self._check_reports()
        except Exception as e:
            self.log_message.emit(f"Auto-scan test envs: {e}", "warning")

    @pyqtSlot(str)
    def load_test_setup(self, tse_path):
        if not self._connected:
            self.log_message.emit("Not connected to CANoe", "error")
            return
        try:
            self.log_message.emit(f"Loading test setup: {tse_path}", "info")
            ts = self._app_com.Configuration.TestSetup
            tse_abs = os.path.abspath(tse_path)

            # Try to add the TSE; if already loaded, scan existing instead
            tenv = None
            try:
                tenv = ts.TestEnvironments.Add(tse_abs)
                tenv = CastTo(tenv, "ITestEnvironment2")
            except Exception as add_err:
                self.log_message.emit(
                    f"TSE may already be loaded, scanning existing environments",
                    "warning",
                )

            self._test_modules.clear()
            if tenv is not None:
                self._collect_modules(tenv)
            else:
                # TSE already loaded - scan all existing environments
                envs = ts.TestEnvironments
                for i in range(1, envs.Count + 1):
                    try:
                        existing = envs.Item(i)
                        existing = CastTo(existing, "ITestEnvironment2")
                        self._collect_modules(existing)
                    except Exception:
                        pass

            # Pump COM messages extensively so CANoe finishes processing
            for _ in range(50):
                self._do_events()

            # Refresh names from COM objects
            self._refresh_module_names()

            info = [(m["name"], m["enabled"]) for m in self._test_modules]
            self.test_modules_found.emit(info)
            self.log_message.emit(
                f"Found {len(self._test_modules)} test module(s) - all enabled in CANoe",
                "info",
            )
            # Auto-scan for existing reports
            self._check_reports()
        except Exception as e:
            self.log_message.emit(f"Load test setup failed: {e}", "error")

    def _collect_modules(self, parent):
        for tm in parent.TestModules:
            # Enable the test module in CANoe (so it gets compiled/built)
            try:
                tm.Enabled = True
            except Exception:
                pass
            ev = DispatchWithEvents(tm, TestEventHandler)
            ev.worker = self
            self._test_modules.append(
                {
                    "name": tm.Name,
                    "enabled": True,
                    "obj": tm,
                    "events": ev,
                    "parent": parent,  # Store parent for re-enumeration
                }
            )
        for folder in parent.Folders:
            self._collect_modules(folder)

    def _refresh_module_names(self):
        """Re-read module names from COM after pumping messages.

        Note: The display name in CANoe may be the CAPL file name,
        but this is not directly accessible via COM API.
        If needed, names can be manually corrected via the UI.
        """
        # Currently just uses the Name property
        # User can manually update test case names if needed
        pass

    @pyqtSlot()
    def start_measurement(self):
        if not self._connected:
            self.log_message.emit("Not connected", "error")
            return
        try:
            if self._meas_com.Running:
                self.log_message.emit("Measurement already running", "warning")
                self.state_changed.emit("running")
                self.measurement_state.emit(True)
                return

            self.state_changed.emit("starting")
            self.log_message.emit("Starting measurement...", "info")
            self.verdict_update.emit("idle")

            self._meas_ev._started = False
            self._meas_com.Start()

            if self._wait_for(lambda: self._meas_ev.is_started, 30):
                self.state_changed.emit("running")
            else:
                self.log_message.emit("Timeout: measurement did not start", "error")
                self.state_changed.emit("connected")
        except Exception as e:
            self.log_message.emit(f"Start failed: {e}", "error")
            self.state_changed.emit("error")

    @pyqtSlot()
    def stop_measurement(self):
        if not self._connected:
            return
        try:
            if not self._meas_com.Running:
                self.log_message.emit("Measurement not running", "warning")
                self.state_changed.emit("connected")
                self.measurement_state.emit(False)
                return

            self.state_changed.emit("stopping")
            self.log_message.emit("Stopping measurement...", "info")

            self._meas_ev._stopped = False
            self._meas_com.Stop()

            if self._wait_for(lambda: self._meas_ev.is_stopped, 30):
                self.state_changed.emit("connected")
            else:
                self.log_message.emit("Timeout: measurement did not stop", "error")
        except Exception as e:
            self.log_message.emit(f"Stop failed: {e}", "error")
            self.state_changed.emit("error")

    @pyqtSlot(list, dict)
    def run_selected_tests(self, selected_names, options):
        """Run tests sequentially with loops, retries, and stop-on-fail."""
        try:
            self.state_changed.emit("testing")
            self.verdict_update.emit("running")

            stop_on_fail = options.get("stop_on_fail", False)
            retry_count = options.get("retry_count", 0)
            loop_count = options.get("loop_count", 1)

            # Build ordered items list respecting drag-drop order
            all_items = self._test_modules + self._test_configs
            items = []
            for name in selected_names:
                for it in all_items:
                    if it["name"] == name:
                        items.append(it)
                        break

            if not items:
                self.log_message.emit(
                    "No test items found matching selection", "warning"
                )
                self.state_changed.emit("running")
                return

            total_per_loop = len(items)
            grand_total = total_per_loop * loop_count
            grand_done = 0
            total_passed = 0
            total_failed = 0
            aborted = False

            self.log_message.emit(
                f"Running {total_per_loop} test(s) x {loop_count} loop(s) = {grand_total} total",
                "info",
            )
            self.progress.emit(0, grand_total)

            for loop_idx in range(loop_count):
                if aborted:
                    break

                if loop_count > 1:
                    self.loop_progress.emit(loop_idx + 1, loop_count)
                    self.log_message.emit(
                        f"--- Loop {loop_idx + 1}/{loop_count} ---", "info"
                    )

                for it in items:
                    if aborted:
                        break

                    final_verdict = VERDICT_NA
                    max_attempts = 1 + retry_count

                    for attempt in range(max_attempts):
                        it["events"].started = False
                        it["events"].stopped = False
                        it["events"].verdict = VERDICT_NA

                        if attempt > 0:
                            self.test_retry.emit(it["name"], attempt, retry_count)
                            self.log_message.emit(
                                f"Retry {attempt}/{retry_count} for '{it['name']}'",
                                "warning",
                            )

                        try:
                            self.log_message.emit(
                                f"Starting test: {it['name']}", "info"
                            )
                            it["obj"].Start()

                            if not self._wait_for(lambda i=it: i["events"].started, 30):
                                self.log_message.emit(
                                    f"Timeout waiting for '{it['name']}' to start",
                                    "warning",
                                )

                            # Wait for completion
                            if not self._wait_for(
                                lambda i=it: i["events"].stopped, 600
                            ):
                                self.log_message.emit(
                                    f"Timeout waiting for '{it['name']}' to finish",
                                    "error",
                                )
                                it["events"].stopped = True
                                it["events"].verdict = VERDICT_NA
                        except Exception as e:
                            self.log_message.emit(
                                f"Failed to run '{it['name']}': {e}", "error"
                            )
                            it["events"].stopped = True
                            it["events"].verdict = VERDICT_NA

                        final_verdict = it["events"].verdict

                        # If event verdict is N/A, try to read from test module object property
                        # (CANoe doesn't always pass verdict through OnStop event)
                        if final_verdict == VERDICT_NA:
                            try:
                                test_obj = it["obj"]
                                if hasattr(test_obj, "Verdict"):
                                    com_verdict = test_obj.Verdict
                                    # CANoe Verdict values: 1=Passed, 2=Failed
                                    if com_verdict == 1:
                                        final_verdict = VERDICT_PASSED
                                    elif com_verdict == 2:
                                        final_verdict = VERDICT_FAILED
                            except Exception:
                                pass

                        # Emit final verdict signal (may override previous OnStop event verdict)
                        if final_verdict != VERDICT_NA:
                            self.test_item_finished.emit(it["name"], final_verdict)

                        if final_verdict == VERDICT_PASSED:
                            break

                    if final_verdict == VERDICT_PASSED:
                        total_passed += 1
                    elif final_verdict == VERDICT_FAILED:
                        total_failed += 1

                    grand_done += 1
                    self.progress.emit(grand_done, grand_total)

                    if stop_on_fail and final_verdict == VERDICT_FAILED:
                        self.log_message.emit(
                            f"Stopping: '{it['name']}' FAILED (stop-on-fail enabled)",
                            "error",
                        )
                        aborted = True
                        break

            self.progress.emit(grand_total, grand_total)
            self.all_tests_complete.emit(total_passed, total_failed)

            if total_failed > 0:
                self.verdict_update.emit("fail")
                self.log_message.emit(
                    f"Done: {total_passed} passed, {total_failed} failed"
                    + (" (aborted)" if aborted else ""),
                    "error",
                )
            else:
                self.verdict_update.emit("pass")
                self.log_message.emit(f"Done: all {grand_total} test(s) passed", "info")

            self.state_changed.emit("running")
            self._check_reports()

        except Exception as e:
            self.log_message.emit(f"Test run error: {e}", "error")
            self.verdict_update.emit("error")
            self.state_changed.emit("running")

    @pyqtSlot(str)
    def get_capl_path(self, test_name):
        """Get the CAPL source path for a test module."""
        try:
            for m in self._test_modules:
                if m["name"] == test_name:
                    path = m["obj"].Path
                    if path:
                        # Construct full path with .can extension from module name
                        full_path = os.path.join(path, test_name + ".can")
                        self.capl_path_found.emit(test_name, full_path)
                        self.log_message.emit(
                            f"CAPL path for '{test_name}': {full_path}", "info"
                        )
                    else:
                        self.log_message.emit(
                            f"No path available for '{test_name}'", "warning"
                        )
                    return
            self.log_message.emit(f"Test module '{test_name}' not found", "warning")
        except Exception as e:
            self.log_message.emit(f"Get CAPL path failed: {e}", "error")

    @pyqtSlot()
    def reload_test_modules(self):
        """Reload all test modules without disconnecting."""
        if not self._connected:
            self.log_message.emit("Not connected", "error")
            return
        try:
            count = 0
            for m in self._test_modules:
                try:
                    obj = m["obj"]
                    # Check if object has Reload method
                    if hasattr(obj, "Reload") and callable(getattr(obj, "Reload")):
                        obj.Reload()
                        count += 1
                    else:
                        self.log_message.emit(
                            f"Reload not supported for '{m['name']}'", "warning"
                        )
                except Exception as e:
                    self.log_message.emit(
                        f"Failed to reload '{m['name']}': {e}", "warning"
                    )
            for _ in range(30):
                self._do_events()
            self.log_message.emit(f"Reloaded {count} test module(s)", "info")
            self._scan_existing_test_envs()
        except Exception as e:
            self.log_message.emit(f"Reload failed: {e}", "error")

    def _check_reports(self):
        """Search for per-test reports and a global latest report."""
        if not self._config_dir:
            return

        # Collect candidate report directories
        report_dirs = []
        for sub in (
            "TestReports",
            "Reports",
            "TestModules_Report",
            "TestConfiguration_Report",
            "Report",
        ):
            d = os.path.join(self._config_dir, sub)
            if os.path.isdir(d):
                report_dirs.append(d)
                self.log_message.emit(f"Scanning report directory: {d}", "info")

        if not report_dirs:
            self.log_message.emit("No report directories found", "warning")
            return

        # Collect ALL report files first
        all_report_files = []
        for d in report_dirs:
            try:
                files = os.listdir(d)
                for f in files:
                    fname_no_ext = os.path.splitext(f)[0]
                    ext_lower = f.lower()

                    # Find all files with valid extension (vtestreport, html, xml)
                    # and containing "report" keyword (flexible: _report, -report, report, etc.)
                    if ("report" in fname_no_ext.lower()) and (
                        ext_lower.endswith(".vtestreport")
                        or ext_lower.endswith(".html")
                        or ext_lower.endswith(".htm")
                        or ext_lower.endswith(".xml")
                    ):
                        full_path = os.path.join(d, f)
                        mtime = os.path.getmtime(full_path)
                        all_report_files.append((fname_no_ext, full_path, mtime))
                        self.log_message.emit(f"Found report file: {f}", "info")
            except Exception as e:
                self.log_message.emit(f"Error scanning {d}: {e}", "warning")

        # Per-test report search - try to match with test names
        all_items = self._test_modules + self._test_configs
        for it in all_items:
            name = it["name"]
            candidates = []

            # Match strategy: Find reports that match the test module name exactly
            for fname_no_ext, full_path, mtime in all_report_files:
                # Remove "_report" or " report" suffix from filename to get base name
                base_name = fname_no_ext
                if base_name.endswith("_report"):
                    base_name = base_name[:-7]  # Remove "_report"
                elif base_name.endswith(" report"):
                    base_name = base_name[:-7]  # Remove " report"
                elif base_name.endswith("-report"):
                    base_name = base_name[:-7]  # Remove "-report"

                # Exact match: base name must equal test name
                if base_name == name:
                    candidates.append((mtime, full_path))

            # Use most recent report if multiple found
            if candidates:
                candidates.sort(reverse=True)  # Sort by mtime, newest first
                report_path = candidates[0][1]
                self.test_report_found.emit(name, report_path)
                self.log_message.emit(
                    f"Mapped report for '{name}': {os.path.basename(report_path)}",
                    "info",
                )
            else:
                self.log_message.emit(
                    f"No matching report found for '{name}'", "warning"
                )

        # Global: emit the most recent report file
        all_reports = []
        for d in report_dirs:
            try:
                for f in os.listdir(d):
                    if f.lower().endswith((".vtestreport", ".html", ".htm", ".xml")):
                        full_path = os.path.join(d, f)
                        mtime = os.path.getmtime(full_path)
                        all_reports.append((mtime, full_path))
            except Exception:
                pass

        if all_reports:
            all_reports.sort(reverse=True)
            self.report_available.emit(all_reports[0][1])

    @pyqtSlot()
    def cleanup(self):
        try:
            if self._write_poll_timer:
                try:
                    if self._write_poll_timer.isActive():
                        self._write_poll_timer.stop()
                except RuntimeError:
                    pass
                try:
                    self._write_poll_timer.deleteLater()
                except RuntimeError:
                    pass
            if self._pump_timer:
                try:
                    if self._pump_timer.isActive():
                        self._pump_timer.stop()
                except RuntimeError:
                    pass
                try:
                    self._pump_timer.deleteLater()
                except RuntimeError:
                    pass
            if self._connected and self._meas_com:
                try:
                    if self._meas_com.Running:
                        self._meas_com.Stop()
                        time.sleep(2)
                except Exception:
                    pass
            if HAS_COM:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
            self._connected = False
        except Exception:
            pass


# ==========================================================================
# Main Window
# ==========================================================================


class MainWindow(QMainWindow):
    """Professional GUI for CANoe Automation Testing."""

    # signals to worker
    sig_connect = pyqtSignal(str)
    sig_load_tse = pyqtSignal(str)
    sig_start = pyqtSignal()
    sig_stop = pyqtSignal()
    sig_run_tests = pyqtSignal(list, dict)
    sig_cleanup = pyqtSignal()
    sig_clear_write = pyqtSignal()
    sig_get_capl_path = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        # Set window icon from logo
        if os.path.isfile(LOGO_PATH):
            icon = QIcon(LOGO_PATH)
            self.setWindowIcon(icon)

        self.setWindowTitle(f"{APP_TITLE}  v{VERSION}")
        self.setMinimumSize(950, 700)
        self.resize(1100, 820)

        self._state = "disconnected"
        self._canoe_ver = ""
        self._report_path = ""
        self._test_start_times = {}
        self._schedule_timer = None
        self._countdown_timer = None
        self._elapsed_timer = None
        self._meas_start_time = None
        self._scheduled_names = []
        self._scheduled_options = {}

        # Persistent settings
        self._settings = QSettings("BLTN", "CANoeTestTool")

        self._build_ui()
        self._build_menu()
        self._build_statusbar()
        self._setup_worker()
        self._update_buttons("disconnected")

        # Restore saved paths
        saved_cfg = self._settings.value("last_cfg_path", "")
        saved_tse = self._settings.value("last_tse_path", "")
        if saved_cfg:
            self.ed_cfg.setText(saved_cfg)
        if saved_tse:
            self.ed_tse.setText(saved_tse)

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(6)
        root.setContentsMargins(8, 8, 8, 8)

        # ---- CONNECTION section ----
        grp_conn = QGroupBox("CONNECTION")
        lo_conn = QVBoxLayout(grp_conn)

        row_cfg = QHBoxLayout()
        row_cfg.addWidget(QLabel("Config (.cfg):"))
        self.ed_cfg = QLineEdit()
        self.ed_cfg.setPlaceholderText("Select CANoe configuration file...")
        row_cfg.addWidget(self.ed_cfg, 1)
        self.btn_browse_cfg = QPushButton("Browse...")
        self.btn_browse_cfg.clicked.connect(self._browse_cfg)
        row_cfg.addWidget(self.btn_browse_cfg)
        self.btn_connect = QPushButton("Connect && Load")
        self.btn_connect.clicked.connect(self._do_connect)
        row_cfg.addWidget(self.btn_connect)
        lo_conn.addLayout(row_cfg)

        row_tse = QHBoxLayout()
        row_tse.addWidget(QLabel("Test Setup (.tse):"))
        self.ed_tse = QLineEdit()
        self.ed_tse.setPlaceholderText("Select test environment file (optional)...")
        row_tse.addWidget(self.ed_tse, 1)
        self.btn_browse_tse = QPushButton("Browse...")
        self.btn_browse_tse.clicked.connect(self._browse_tse)
        row_tse.addWidget(self.btn_browse_tse)
        self.btn_load_tse = QPushButton("Load TSE")
        self.btn_load_tse.clicked.connect(self._do_load_tse)
        row_tse.addWidget(self.btn_load_tse)
        lo_conn.addLayout(row_tse)

        self.lbl_conn = QLabel("Status: Disconnected")
        self.lbl_conn.setStyleSheet("color:#6c6c6c; font-weight:bold;")
        lo_conn.addWidget(self.lbl_conn)

        root.addWidget(grp_conn)

        # ---- middle: vertical splitter (tree+ctrl | console) ----
        v_split = QSplitter(Qt.Orientation.Vertical)

        # -- horizontal splitter: test tree | control panel --
        h_split = QSplitter(Qt.Orientation.Horizontal)

        # left: test tree
        grp_tree = QGroupBox("TEST SELECTION")
        lo_tree = QVBoxLayout(grp_tree)
        self.tree = IntraGroupTreeWidget()
        self.tree.setHeaderLabels(["Test Item", "Status", "Duration", "Report"])
        self.tree.setColumnCount(4)
        hdr = self.tree.header()
        hdr.setStretchLastSection(False)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        self.tree.setColumnWidth(3, 150)
        self.tree.setAlternatingRowColors(True)
        self.tree.itemDoubleClicked.connect(self._on_tree_double_clicked)
        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._show_tree_context_menu)
        lo_tree.addWidget(self.tree)

        row_sel = QHBoxLayout()
        b_all = QPushButton("Select All")
        b_all.clicked.connect(self._select_all)
        b_none = QPushButton("Deselect All")
        b_none.clicked.connect(self._deselect_all)
        self.btn_edit_capl = QPushButton("Edit CAPL")
        self.btn_edit_capl.clicked.connect(self._do_edit_capl)
        self.btn_delete = QPushButton("Delete Selected")
        self.btn_delete.clicked.connect(self._do_delete_selected)
        row_sel.addWidget(b_all)
        row_sel.addWidget(b_none)
        row_sel.addWidget(self.btn_edit_capl)
        row_sel.addWidget(self.btn_delete)
        row_sel.addStretch()
        lo_tree.addLayout(row_sel)

        h_split.addWidget(grp_tree)

        # right: control & status
        grp_ctrl = QGroupBox("CONTROL && STATUS")
        lo_ctrl = QVBoxLayout(grp_ctrl)
        lo_ctrl.setSpacing(10)

        self.btn_start = QPushButton("START MEASUREMENT")
        self.btn_start.setObjectName("btn_start")
        self.btn_start.setMinimumHeight(44)
        self.btn_start.clicked.connect(self._do_start)
        lo_ctrl.addWidget(self.btn_start)

        self.btn_stop = QPushButton("STOP MEASUREMENT")
        self.btn_stop.setObjectName("btn_stop")
        self.btn_stop.setMinimumHeight(44)
        self.btn_stop.clicked.connect(self._do_stop)
        lo_ctrl.addWidget(self.btn_stop)

        lo_ctrl.addSpacing(4)

        row_options = QHBoxLayout()
        self.chk_stop_on_fail = QCheckBox("Stop on first failure")
        row_options.addWidget(self.chk_stop_on_fail)

        row_options.addWidget(QLabel("Retry failed:"))
        self.spn_retry = QSpinBox()
        self.spn_retry.setRange(0, 10)
        self.spn_retry.setValue(0)
        self.spn_retry.setSuffix(" times")
        row_options.addWidget(self.spn_retry)

        row_options.addWidget(QLabel("Loop count:"))
        self.spn_loop = QSpinBox()
        self.spn_loop.setRange(1, 1000)
        self.spn_loop.setValue(1)
        self.spn_loop.setSuffix(" time(s)")
        row_options.addWidget(self.spn_loop)
        row_options.addStretch()
        lo_ctrl.addLayout(row_options)

        lo_ctrl.addSpacing(4)

        self.btn_run = QPushButton("RUN SELECTED TESTS")
        self.btn_run.setObjectName("btn_run")
        self.btn_run.setMinimumHeight(44)
        self.btn_run.clicked.connect(self._do_run_tests)
        lo_ctrl.addWidget(self.btn_run)

        lo_ctrl.addSpacing(8)

        row_meas = QHBoxLayout()
        self.lbl_meas = QLabel("Measurement: Stopped")
        self.lbl_meas.setStyleSheet("font-size:10pt; color:#6c6c6c;")
        row_meas.addWidget(self.lbl_meas)

        self.lbl_elapsed = QLabel("")
        self.lbl_elapsed.setStyleSheet("font-size:8pt; color:#9cdcfe; margin-left:4px;")
        row_meas.addWidget(self.lbl_elapsed)
        row_meas.addStretch()
        lo_ctrl.addLayout(row_meas)

        lo_ctrl.addSpacing(8)

        # Date, Time, Schedule, and Buttons all in one row
        row_datetime = QHBoxLayout()
        lbl_sched_label = QLabel("Schedule:")
        lbl_sched_label.setStyleSheet("color:#4a90e2; font-weight:bold; font-size:9pt;")
        row_datetime.addWidget(lbl_sched_label)
        row_datetime.addWidget(QLabel("Date:"))
        self.dt_schedule_date = QDateTimeEdit()
        self.dt_schedule_date.setDisplayFormat("yyyy-MM-dd")
        self.dt_schedule_date.setDateTime(QDateTime.currentDateTime().addSecs(300))
        self.dt_schedule_date.setCalendarPopup(True)
        row_datetime.addWidget(self.dt_schedule_date, 1)

        row_datetime.addWidget(QLabel("Time:"))
        self.dt_schedule_time = QDateTimeEdit()
        self.dt_schedule_time.setDisplayFormat("HH:mm:ss")
        self.dt_schedule_time.setDateTime(QDateTime.currentDateTime().addSecs(300))
        row_datetime.addWidget(self.dt_schedule_time, 1)

        self.btn_schedule = QPushButton("Set")
        self.btn_schedule.clicked.connect(self._do_schedule)
        row_datetime.addWidget(self.btn_schedule)
        self.btn_cancel_schedule = QPushButton("Cancel")
        self.btn_cancel_schedule.clicked.connect(self._cancel_schedule)
        self.btn_cancel_schedule.setEnabled(False)
        row_datetime.addWidget(self.btn_cancel_schedule)
        lo_ctrl.addLayout(row_datetime)

        self.lbl_schedule = QLabel("")
        self.lbl_schedule.setStyleSheet("color:#6c6c6c; font-size:9pt;")
        lo_ctrl.addWidget(self.lbl_schedule)

        lo_ctrl.addStretch()
        h_split.addWidget(grp_ctrl)
        h_split.setSizes([550, 300])

        v_split.addWidget(h_split)

        # ---- console + write window side by side ----
        console_split = QSplitter(Qt.Orientation.Horizontal)

        # left: console log
        grp_con = QGroupBox("CONSOLE LOG")
        lo_con = QVBoxLayout(grp_con)
        self.console = QTextEdit()
        self.console.setReadOnly(True)
        lo_con.addWidget(self.console)

        row_con = QHBoxLayout()
        b_clear = QPushButton("Clear")
        b_clear.clicked.connect(self.console.clear)
        row_con.addWidget(b_clear)
        row_con.addStretch()
        lo_con.addLayout(row_con)

        console_split.addWidget(grp_con)

        # right: write window (mirrors CANoe Write Window output)
        grp_write = QGroupBox("WRITE WINDOW")
        lo_write = QVBoxLayout(grp_write)
        self.write_win = QTextEdit()
        self.write_win.setReadOnly(True)
        lo_write.addWidget(self.write_win)

        row_wr = QHBoxLayout()
        b_clear_wr = QPushButton("Clear")
        b_clear_wr.clicked.connect(self._clear_write_window)
        row_wr.addWidget(b_clear_wr)
        row_wr.addStretch()
        lo_write.addLayout(row_wr)

        console_split.addWidget(grp_write)
        console_split.setSizes([500, 500])

        v_split.addWidget(console_split)
        v_split.setSizes([420, 250])

        root.addWidget(v_split, 1)

        # ---- bottom: progress + export ----
        row_bot = QHBoxLayout()
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setFormat("Idle")
        row_bot.addWidget(self.progress, 1)

        self.btn_export = QPushButton("Export to Excel")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._do_export_excel)
        row_bot.addWidget(self.btn_export)

        root.addLayout(row_bot)

    def _build_menu(self):
        mb = self.menuBar()

        m_file = mb.addMenu("File")
        a = QAction("Open Config...", self)
        a.setShortcut("Ctrl+O")
        a.triggered.connect(self._browse_cfg)
        m_file.addAction(a)
        m_file.addSeparator()
        a = QAction("Exit", self)
        a.setShortcut("Ctrl+Q")
        a.triggered.connect(self.close)
        m_file.addAction(a)

        m_tools = mb.addMenu("Tools")
        a = QAction("Clear Console", self)
        a.triggered.connect(self.console.clear)
        m_tools.addAction(a)
        a = QAction("Clear Write Window", self)
        a.triggered.connect(self._clear_write_window)
        m_tools.addAction(a)
        m_tools.addSeparator()
        a = QAction("Export Results to Excel...", self)
        a.triggered.connect(self._do_export_excel)
        m_tools.addAction(a)

        m_help = mb.addMenu("Help")
        a = QAction("About", self)
        a.triggered.connect(self._show_about)
        m_help.addAction(a)

    def _build_statusbar(self):
        sb = self.statusBar()
        self.sb_ver = QLabel("  CANoe: N/A")
        self.sb_state = QLabel("Disconnected  ")
        sb.addWidget(self.sb_ver, 1)
        sb.addPermanentWidget(self.sb_state)

    # ------------------------------------------------------------------
    # Worker thread setup
    # ------------------------------------------------------------------

    def _setup_worker(self):
        self._thread = QThread()
        self._worker = CANoeWorker()
        self._worker.moveToThread(self._thread)

        # thread start -> worker init
        self._thread.started.connect(self._worker.initialize)

        # UI -> Worker
        self.sig_connect.connect(self._worker.connect_and_load)
        self.sig_load_tse.connect(self._worker.load_test_setup)
        self.sig_start.connect(self._worker.start_measurement)
        self.sig_stop.connect(self._worker.stop_measurement)
        self.sig_run_tests.connect(self._worker.run_selected_tests)
        self.sig_cleanup.connect(self._worker.cleanup)
        self.sig_clear_write.connect(self._worker.reset_write_cache)
        self.sig_get_capl_path.connect(self._worker.get_capl_path)

        # Worker -> UI
        self._worker.log_message.connect(self._on_log)
        self._worker.state_changed.connect(self._update_buttons)
        self._worker.canoe_connected.connect(self._on_connected)
        self._worker.canoe_quit.connect(lambda: self._update_buttons("disconnected"))
        self._worker.measurement_state.connect(self._on_meas_state)
        self._worker.test_modules_found.connect(self._populate_modules)
        self._worker.test_configs_found.connect(self._populate_configs)
        self._worker.test_item_started.connect(self._tree_set_running)
        self._worker.test_item_finished.connect(self._tree_set_verdict)
        self._worker.all_tests_complete.connect(self._on_tests_done)
        self._worker.progress.connect(self._on_progress)
        self._worker.report_available.connect(self._on_report)
        self._worker.test_report_found.connect(self._on_test_report)
        self._worker.write_window_text.connect(self._on_write_window)
        self._worker.capl_path_found.connect(self._on_capl_path)
        self._worker.test_retry.connect(self._on_test_retry)
        self._worker.loop_progress.connect(self._on_loop_progress)

        self._thread.start()

    # ------------------------------------------------------------------
    # UI event handlers (user clicks)
    # ------------------------------------------------------------------

    def _browse_cfg(self):
        p, _ = QFileDialog.getOpenFileName(
            self,
            "Select CANoe Config",
            self.ed_cfg.text() or "",
            "CANoe Config (*.cfg);;All Files (*)",
        )
        if p:
            self.ed_cfg.setText(p)

    def _browse_tse(self):
        start = self.ed_tse.text() or os.path.dirname(self.ed_cfg.text()) or ""
        p, _ = QFileDialog.getOpenFileName(
            self, "Select Test Setup", start, "Test Setup (*.tse);;All Files (*)"
        )
        if p:
            self.ed_tse.setText(p)

    def _do_connect(self):
        p = self.ed_cfg.text().strip()
        if not p:
            QMessageBox.warning(self, "Warning", "Please select a .cfg file first.")
            return
        if not os.path.isfile(p):
            QMessageBox.warning(self, "Warning", f"File not found:\n{p}")
            return
        self._save_paths()
        self.sig_connect.emit(p)

    def _do_load_tse(self):
        p = self.ed_tse.text().strip()
        if not p:
            QMessageBox.warning(self, "Warning", "Please select a .tse file first.")
            return
        self._save_paths()
        self.sig_load_tse.emit(p)

    def _do_start(self):
        self.sig_start.emit()

    def _do_stop(self):
        self.sig_stop.emit()

    def _do_run_tests(self):
        sel = self._get_checked()
        if not sel:
            QMessageBox.information(
                self, "Info", "No tests selected. Check items in the tree first."
            )
            return
        options = {
            "stop_on_fail": self.chk_stop_on_fail.isChecked(),
            "retry_count": self.spn_retry.value(),
            "loop_count": self.spn_loop.value(),
        }
        self.sig_run_tests.emit(sel, options)

    def _get_checked(self):
        names = []
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            grp = root.child(i)
            for j in range(grp.childCount()):
                item = grp.child(j)
                if item.checkState(0) == Qt.CheckState.Checked:
                    names.append(item.text(0))
        return names

    def _select_all(self):
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            g = root.child(i)
            for j in range(g.childCount()):
                g.child(j).setCheckState(0, Qt.CheckState.Checked)

    def _deselect_all(self):
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            g = root.child(i)
            for j in range(g.childCount()):
                g.child(j).setCheckState(0, Qt.CheckState.Unchecked)

    def _on_tree_double_clicked(self, item, column):
        """Handle tree item double-click - open report if Report column."""
        if column == 3:
            path = item.data(3, Qt.ItemDataRole.UserRole)
            if path and os.path.isfile(path):
                try:
                    os.startfile(path)
                    self._on_log(f"Opening report: {os.path.basename(path)}", "info")
                except Exception as e:
                    QMessageBox.warning(
                        self, "Error", f"Cannot open report file:\n{path}\n\nError: {e}"
                    )

    def _show_tree_context_menu(self, pos):
        """Right-click context menu on tree items."""
        item = self.tree.itemAt(pos)
        if item is None or item.parent() is None:
            return
        menu = QMenu(self)
        a_edit = menu.addAction("Edit CAPL Source...")
        a_edit.triggered.connect(lambda: self._do_edit_capl_for(item.text(0)))
        menu.exec(self.tree.viewport().mapToGlobal(pos))

    def _do_edit_capl(self):
        """Edit CAPL for the currently selected tree item."""
        item = self.tree.currentItem()
        if item is None or item.parent() is None:
            QMessageBox.information(self, "Info", "Select a test module first.")
            return
        self._do_edit_capl_for(item.text(0))

    def _do_edit_capl_for(self, name):
        self.sig_get_capl_path.emit(name)

    def _do_delete_selected(self):
        """Delete selected test items from tree."""
        selected = self._get_checked()
        if not selected:
            QMessageBox.information(self, "Info", "No tests selected to delete.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete {len(selected)} selected test(s)?\n\nThis cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        # Delete from tree
        root = self.tree.invisibleRootItem()
        deleted_count = 0
        for i in range(root.childCount()):
            grp = root.child(i)
            j = 0
            while j < grp.childCount():
                item = grp.child(j)
                if item.text(0) in selected:
                    grp.removeChild(item)
                    deleted_count += 1
                else:
                    j += 1

        self._on_log(f"Deleted {deleted_count} test(s)", "info")

    def _on_capl_path(self, name, path):
        """Open the CAPL source file using CAPLBrowser."""
        if path and os.path.isfile(path):
            try:
                capl_browser = (
                    r"C:\Program Files\Vector CANoe 17\Exec64\CAPLBrowser.exe"
                )
                if os.path.isfile(capl_browser):
                    subprocess.Popen([capl_browser, path])
                    self._on_log(f"Opening CAPL in CAPLBrowser: {path}", "info")
                else:
                    # Fallback to default editor if CAPLBrowser not found
                    os.startfile(path)
                    self._on_log(f"Opening CAPL with default editor: {path}", "info")
            except Exception as e:
                QMessageBox.warning(
                    self, "Error", f"Cannot open CAPL file:\n{path}\n\nError: {e}"
                )
        else:
            QMessageBox.information(
                self, "Info", f"CAPL path not found for '{name}':\n{path}"
            )

    def _do_export_excel(self):
        """Export test results to .xlsx file - only selected (checked) tests."""
        if not HAS_OPENPYXL:
            QMessageBox.warning(
                self,
                "Warning",
                "openpyxl is required for Excel export.\n\npip install openpyxl",
            )
            return

        # Only export checked tests
        rows = []
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            grp = root.child(i)
            group_name = grp.text(0)
            for j in range(grp.childCount()):
                it = grp.child(j)
                # Only include checked items
                if it.checkState(0) == Qt.CheckState.Checked:
                    report_path = it.data(3, Qt.ItemDataRole.UserRole) or ""
                    rows.append(
                        {
                            "group": group_name,
                            "name": it.text(0),
                            "status": it.text(1),
                            "duration": it.text(2),
                            "report": it.text(3),
                            "report_path": report_path,
                        }
                    )
        if not rows:
            QMessageBox.information(self, "Info", "No selected tests to export. Please check tests in the tree first.")
            return
        default_name = (
            f"TestResults_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results",
            default_name,
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Results"
            title_font = XlFont(name="Segoe UI", size=14, bold=True, color="0E639C")
            header_font = XlFont(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="0E639C", end_color="0E639C", fill_type="solid"
            )
            pass_fill = PatternFill(
                start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
            )
            fail_fill = PatternFill(
                start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
            )
            pass_font = XlFont(name="Segoe UI", size=10, bold=True, color="2E7D32")
            fail_font = XlFont(name="Segoe UI", size=10, bold=True, color="C62828")
            normal_font = XlFont(name="Segoe UI", size=10)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Title and metadata
            ws.merge_cells("A1:G1")
            ws["A1"] = f"{APP_TITLE} - Test Results (Selected Tests Only)"
            ws["A1"].font = title_font
            ws["A2"] = f"Export Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A2"].font = normal_font
            ws["A3"] = f"CANoe Version: {self._canoe_ver or 'N/A'}"
            ws["A3"].font = normal_font
            ws["A4"] = f"Config: {self.ed_cfg.text()}"
            ws["A4"].font = normal_font
            ws["A5"] = f"Test Setup: {self.ed_tse.text() or 'N/A'}"
            ws["A5"].font = normal_font

            # Test execution settings
            ws["A6"] = f"Settings: Stop on Fail={self.chk_stop_on_fail.isChecked()}, Retry={self.spn_retry.value()}, Loop={self.spn_loop.value()}"
            ws["A6"].font = normal_font

            row_num = 8
            headers = ["#", "Group", "Test Name", "Status", "Duration", "Report File", "Report Path"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            passed_count = 0
            failed_count = 0
            na_count = 0

            for idx, r in enumerate(rows, 1):
                row_num += 1
                ws.cell(row=row_num, column=1, value=idx).border = thin_border
                ws.cell(row=row_num, column=2, value=r["group"]).border = thin_border
                ws.cell(row=row_num, column=3, value=r["name"]).border = thin_border

                status_cell = ws.cell(row=row_num, column=4, value=r["status"])
                status_cell.border = thin_border
                status_cell.alignment = Alignment(horizontal="center")
                if r["status"] == "PASSED":
                    status_cell.font = pass_font
                    status_cell.fill = pass_fill
                    passed_count += 1
                elif r["status"] == "FAILED":
                    status_cell.font = fail_font
                    status_cell.fill = fail_fill
                    failed_count += 1
                else:
                    status_cell.font = normal_font
                    na_count += 1

                ws.cell(row=row_num, column=5, value=r["duration"]).border = thin_border
                ws.cell(row=row_num, column=6, value=r["report"]).border = thin_border

                # Add hyperlink to report path if available
                report_path_cell = ws.cell(row=row_num, column=7, value=r["report_path"])
                report_path_cell.border = thin_border
                if r["report_path"] and os.path.isfile(r["report_path"]):
                    report_path_cell.hyperlink = r["report_path"]
                    report_path_cell.font = XlFont(name="Segoe UI", size=10, color="0563C1", underline="single")

            # Summary section
            row_num += 2
            ws.cell(row=row_num, column=1, value="Summary:").font = XlFont(
                name="Segoe UI", size=11, bold=True
            )
            ws.cell(row=row_num, column=2, value=f"Total Selected: {len(rows)}").font = normal_font
            ws.cell(row=row_num, column=3, value=f"Passed: {passed_count}").font = pass_font
            ws.cell(row=row_num, column=4, value=f"Failed: {failed_count}").font = fail_font
            ws.cell(row=row_num, column=5, value=f"Not Run: {na_count}").font = normal_font

            # Calculate pass rate
            if passed_count + failed_count > 0:
                pass_rate = (passed_count / (passed_count + failed_count)) * 100
                ws.cell(row=row_num, column=6, value=f"Pass Rate: {pass_rate:.1f}%").font = normal_font

            # Adjust column widths
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 35
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 25
            ws.column_dimensions["G"].width = 50

            wb.save(path)
            self._on_log(f"Results exported to: {path}", "info")
            reply = QMessageBox.question(
                self,
                "Export Complete",
                f"{len(rows)} selected test(s) exported to:\n{path}\n\nOpen the file now?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                os.startfile(path)
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export results:\n{e}")

    def _do_schedule(self):
        """Set a one-time scheduled test execution."""
        sel = self._get_checked()
        if not sel:
            QMessageBox.information(
                self, "Info", "No tests selected. Check items in the tree first."
            )
            return

        # Combine date and time from separate widgets
        date = self.dt_schedule_date.date()
        time = self.dt_schedule_time.time()
        target = QDateTime(date, time)

        now = QDateTime.currentDateTime()
        diff_ms = now.msecsTo(target)
        if diff_ms <= 0:
            QMessageBox.warning(
                self, "Warning", "Scheduled time must be in the future."
            )
            return
        self._scheduled_names = sel
        self._scheduled_options = {
            "stop_on_fail": self.chk_stop_on_fail.isChecked(),
            "retry_count": self.spn_retry.value(),
            "loop_count": self.spn_loop.value(),
        }
        self._schedule_timer = QTimer(self)
        self._schedule_timer.setSingleShot(True)
        self._schedule_timer.timeout.connect(self._on_scheduled_fire)
        self._schedule_timer.start(int(diff_ms))
        self._countdown_timer = QTimer(self)
        self._countdown_timer.timeout.connect(self._update_countdown)
        self._countdown_timer.start(1000)
        self.btn_schedule.setEnabled(False)
        self.btn_cancel_schedule.setEnabled(True)
        self.dt_schedule_date.setEnabled(False)
        self.dt_schedule_time.setEnabled(False)
        self._update_countdown()
        self._on_log(
            f"Tests scheduled for {target.toString('yyyy-MM-dd HH:mm:ss')}", "info"
        )

    def _update_countdown(self):
        if not self._schedule_timer or not self._schedule_timer.isActive():
            self.lbl_schedule.setText("")
            return
        remaining = self._schedule_timer.remainingTime() // 1000
        if remaining < 0:
            remaining = 0
        h = remaining // 3600
        m = (remaining % 3600) // 60
        s = remaining % 60
        self.lbl_schedule.setText(f"Starts in: {h:02d}:{m:02d}:{s:02d}")
        self.lbl_schedule.setStyleSheet(
            "color:#dcdcaa; font-size:9pt; font-weight:bold;"
        )

    def _cancel_schedule(self):
        if self._schedule_timer:
            self._schedule_timer.stop()
            self._schedule_timer = None
        if self._countdown_timer:
            self._countdown_timer.stop()
            self._countdown_timer = None
        self.lbl_schedule.setText("Schedule cancelled")
        self.lbl_schedule.setStyleSheet("color:#6c6c6c; font-size:9pt;")
        self.btn_schedule.setEnabled(True)
        self.btn_cancel_schedule.setEnabled(False)
        self.dt_schedule_date.setEnabled(True)
        self.dt_schedule_time.setEnabled(True)
        self._on_log("Scheduled execution cancelled", "warning")

    def _on_scheduled_fire(self):
        if self._countdown_timer:
            self._countdown_timer.stop()
            self._countdown_timer = None
        self._schedule_timer = None
        self.lbl_schedule.setText("")
        self.btn_schedule.setEnabled(True)
        self.btn_cancel_schedule.setEnabled(False)
        self.dt_schedule_date.setEnabled(True)
        self.dt_schedule_time.setEnabled(True)
        self._on_log("Scheduled execution triggered", "info")

        # Check if measurement is already running
        if self._state != "running":
            # Auto-start measurement first
            self._on_log("Auto-starting measurement for scheduled tests", "info")
            self.sig_start.emit()
            # Emit run_tests signal after a brief delay to ensure measurement is starting
            QTimer.singleShot(500, lambda: self.sig_run_tests.emit(self._scheduled_names, self._scheduled_options))
        else:
            # Measurement already running, run tests immediately
            self.sig_run_tests.emit(self._scheduled_names, self._scheduled_options)

    def _on_test_retry(self, name, attempt, max_retries):
        self._tree_status(name, f"Retry {attempt}/{max_retries}", "#dcdcaa")

    def _on_loop_progress(self, current, total):
        self.sb_state.setText(f"  Loop {current}/{total} - Running Tests...  ")

    def _clear_write_window(self):
        """Clear the Write Window display and reset worker's cache."""
        self.write_win.clear()
        self.sig_clear_write.emit()

    def _save_paths(self):
        """Persist current file paths to QSettings."""
        self._settings.setValue("last_cfg_path", self.ed_cfg.text().strip())
        self._settings.setValue("last_tse_path", self.ed_tse.text().strip())

    def _show_about(self):
        QMessageBox.about(
            self,
            "About",
            f"<b>{APP_TITLE}</b><br>Version {VERSION}<br><br>"
            "Professional CANoe Automation Test GUI<br>"
            "Built with PyQt6 + CANoe COM API<br><br>"
            "Architecture: UI Thread  &harr;  Worker Thread  &harr;  "
            "CANoe COM Server",
        )

    # ------------------------------------------------------------------
    # Worker signal handlers
    # ------------------------------------------------------------------

    def _on_log(self, msg, level):
        ts = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        c = {"info": "#d4d4d4", "warning": "#dcdcaa", "error": "#f44747"}.get(
            level, "#d4d4d4"
        )
        pfx = {"warning": "WARNING: ", "error": "ERROR: "}.get(level, "")
        html = (
            f'<span style="color:#6a9955">[{ts}]</span> '
            f'<span style="color:{c}">{pfx}{msg}</span>'
        )
        self.console.append(html)
        cur = self.console.textCursor()
        cur.movePosition(QTextCursor.MoveOperation.End)
        self.console.setTextCursor(cur)

    def _on_connected(self, ver):
        """Handle CANoe connection success."""
        self._canoe_ver = ver
        self.lbl_conn.setText(f"Status: Connected  (CANoe {ver})")
        self.lbl_conn.setStyleSheet("color:#4caf50; font-weight:bold;")
        self.sb_ver.setText(f"  CANoe: {ver}")

    def _update_elapsed(self):
        """Update elapsed time display."""
        if not self._meas_start_time:
            return
        elapsed = datetime.datetime.now() - self._meas_start_time
        total_seconds = int(elapsed.total_seconds())
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        self.lbl_elapsed.setText(f"Elapsed: {minutes:02d}:{seconds:02d}")

    def _on_meas_state(self, running):
        """Sync UI with measurement state (handles external CANoe changes)."""
        if running:
            self.lbl_meas.setText("Measurement: RUNNING")
            self.lbl_meas.setStyleSheet(
                "font-size:10pt; color:#4caf50; font-weight:bold;"
            )
            # Start elapsed time tracking
            self._meas_start_time = datetime.datetime.now()
            if not self._elapsed_timer:
                self._elapsed_timer = QTimer(self)
                self._elapsed_timer.timeout.connect(self._update_elapsed)
            self._elapsed_timer.start(1000)
            self._update_elapsed()
            # Update buttons unless the tool itself is managing the transition
            if self._state not in ("starting", "testing"):
                self._update_buttons("running")
        else:
            self.lbl_meas.setText("Measurement: STOPPED")
            self.lbl_meas.setStyleSheet(
                "font-size:10pt; color:#6c6c6c; font-weight:bold;"
            )
            # Stop elapsed time tracking
            if self._elapsed_timer:
                self._elapsed_timer.stop()
            self.lbl_elapsed.setText("")
            self._meas_start_time = None
            if self._state not in ("stopping", "testing", "connecting"):
                self._update_buttons("connected")

    def _populate_modules(self, modules):
        grp = self._get_or_create_group("Test Modules")
        while grp.childCount():
            grp.removeChild(grp.child(0))
        for name, enabled in modules:
            ch = QTreeWidgetItem(grp, [name, "Idle", "", ""])
            ch.setCheckState(0, Qt.CheckState.Unchecked)
            ch.setForeground(1, QColor("#6c6c6c"))

    def _populate_configs(self, configs):
        grp = self._get_or_create_group("Test Configurations")
        while grp.childCount():
            grp.removeChild(grp.child(0))
        for name, enabled in configs:
            ch = QTreeWidgetItem(grp, [name, "Idle", "", ""])
            ch.setCheckState(
                0, Qt.CheckState.Checked if enabled else Qt.CheckState.Unchecked
            )
            ch.setForeground(1, QColor("#6c6c6c"))

    def _get_or_create_group(self, label):
        for i in range(self.tree.topLevelItemCount()):
            if self.tree.topLevelItem(i).text(0) == label:
                return self.tree.topLevelItem(i)
        item = QTreeWidgetItem(self.tree, [label, "", "", ""])
        item.setExpanded(True)
        f = item.font(0)
        f.setBold(True)
        item.setFont(0, f)
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsDragEnabled)
        return item

    def _tree_set_running(self, name):
        self._test_start_times[name] = time.time()
        self._tree_status(name, "Running", "#569cd6")

    def _tree_set_verdict(self, name, verdict):
        m = {0: ("N/A", "#6c6c6c"), 1: ("PASSED", "#4caf50"), 2: ("FAILED", "#f44336")}
        txt, col = m.get(verdict, ("?", "#ff9800"))
        self._tree_status(name, txt, col)
        if name in self._test_start_times:
            elapsed = time.time() - self._test_start_times[name]
            self._tree_set_column(name, 2, self._format_duration(elapsed), "#9cdcfe")
            del self._test_start_times[name]

    def _format_duration(self, seconds):
        if seconds < 60:
            return f"{seconds:.1f}s"
        elif seconds < 3600:
            m = int(seconds) // 60
            s = int(seconds) % 60
            return f"{m}m {s}s"
        else:
            h = int(seconds) // 3600
            m = (int(seconds) % 3600) // 60
            return f"{h}h {m}m"

    def _tree_set_column(self, name, col, text, color):
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            g = root.child(i)
            for j in range(g.childCount()):
                it = g.child(j)
                if it.text(0) == name:
                    it.setText(col, text)
                    it.setForeground(col, QColor(color))
                    return

    def _tree_status(self, name, status, color):
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            g = root.child(i)
            for j in range(g.childCount()):
                it = g.child(j)
                if it.text(0) == name:
                    it.setText(1, status)
                    it.setForeground(1, QColor(color))
                    return

    def _on_tests_done(self, passed, failed):
        self._on_log(
            f"Summary: {passed} Passed, {failed} Failed",
            "info" if failed == 0 else "error",
        )
        self.btn_export.setEnabled(True)

    def _on_progress(self, done, total):
        if total > 0:
            pct = int(done / total * 100)
            self.progress.setMaximum(100)
            self.progress.setValue(pct)
            self.progress.setFormat(f"{done}/{total}  ({pct}%)")
        else:
            self.progress.setValue(0)
            self.progress.setFormat("Idle")

    def _on_report(self, path):
        self._report_path = path
        self._on_log(f"Report: {path}", "info")

    def _on_test_report(self, name, path):
        """Set the Report column (col 3) for a specific test item."""
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            g = root.child(i)
            for j in range(g.childCount()):
                it = g.child(j)
                if it.text(0) == name:
                    report_filename = os.path.basename(path)
                    it.setText(3, report_filename)
                    it.setForeground(3, QColor("#4caf50"))
                    it.setData(3, Qt.ItemDataRole.UserRole, path)
                    font = it.font(3)
                    font.setBold(True)
                    it.setFont(3, font)
                    it.setToolTip(3, f"Double-click to open: {report_filename}")
                    return

    def _on_write_window(self, text):
        """Display new Write Window content from CANoe."""
        for line in text.splitlines():
            stripped = line.rstrip()
            if stripped:
                self.write_win.append(f'<span style="color:#9cdcfe">{stripped}</span>')
        cur = self.write_win.textCursor()
        cur.movePosition(QTextCursor.MoveOperation.End)
        self.write_win.setTextCursor(cur)

    # ------------------------------------------------------------------
    # LED indicator helper  (removed)
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # State-driven button enable / disable
    # ------------------------------------------------------------------

    def _update_buttons(self, state):
        self._state = state
        conn = state == "connected"
        run = state == "running"
        testing = state == "testing"
        busy = state in ("connecting", "starting", "stopping", "testing")

        # connection area
        self.btn_browse_cfg.setEnabled(not busy and not run)
        self.btn_connect.setEnabled(not busy and not run)
        self.btn_browse_tse.setEnabled(conn or run or testing)
        self.btn_load_tse.setEnabled(conn or run or testing)

        # control area
        self.btn_start.setEnabled(conn)
        self.btn_stop.setEnabled(run or testing)  # Allow stop during testing
        self.btn_run.setEnabled(run and not testing)  # Disable run during testing

        # new controls
        self.chk_stop_on_fail.setEnabled(not busy)
        self.spn_retry.setEnabled(not busy)
        self.spn_loop.setEnabled(not busy)
        self.btn_edit_capl.setEnabled(conn or run)
        self.btn_delete.setEnabled(not busy)
        self.tree.setDragEnabled(not busy)

        # schedule - don't override if a schedule is active
        schedule_active = (
            self._schedule_timer is not None and self._schedule_timer.isActive()
        )
        if not schedule_active:
            self.btn_schedule.setEnabled((conn or run) and not busy)
            self.dt_schedule_date.setEnabled((conn or run) and not busy)
            self.dt_schedule_time.setEnabled((conn or run) and not busy)

        labels = {
            "disconnected": "Disconnected",
            "connecting": "Connecting...",
            "connected": "Connected",
            "starting": "Starting...",
            "running": "Measurement Running",
            "stopping": "Stopping...",
            "testing": "Running Tests...",
            "error": "Error",
        }
        self.sb_state.setText(f"  {labels.get(state, state)}  ")

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------

    def closeEvent(self, event):
        self._save_paths()
        # Don't process signals during cleanup to avoid cross-thread issues
        self.blockSignals(True)
        # Stop timers safely
        try:
            if self._schedule_timer and self._schedule_timer.isActive():
                self._schedule_timer.stop()
        except RuntimeError:
            pass
        try:
            if self._countdown_timer and self._countdown_timer.isActive():
                self._countdown_timer.stop()
        except RuntimeError:
            pass
        try:
            if self._elapsed_timer and self._elapsed_timer.isActive():
                self._elapsed_timer.stop()
        except RuntimeError:
            pass
        self.sig_cleanup.emit()
        self._thread.quit()
        if not self._thread.wait(5000):
            self._thread.terminate()
        event.accept()


# ==========================================================================
# Entry point
# ==========================================================================


def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(DARK_STYLE)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
